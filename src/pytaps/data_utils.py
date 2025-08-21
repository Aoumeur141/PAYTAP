# pytap/data_utils.py

"""
Data Utilities for common Pandas DataFrame operations.
"""

import logging
import pandas as pd

from datetime import datetime
import logging # <-- NEW: Import logging

import pandas as pd
from pathlib import Path
from typing import List, Optional, Any, Union, Tuple # MODIFIED: Added Tuple
import openpyxl # This is crucial for openpyxl.worksheet.worksheet.Worksheet

import openpyxl.worksheet.worksheet # For type hinting the Worksheet object
from openpyxl import Workbook # Added
from openpyxl.utils.dataframe import dataframe_to_rows # Added
import os

import pdbufr # <--- Make sure pdbufr is imported here for the new function
logger = logging.getLogger(__name__) # Module-level logger

def select_existing_columns(df: pd.DataFrame, desired_columns: List[str], logger_instance: Optional[logging.Logger] = None) -> pd.DataFrame:
    """
    Selects a subset of columns from a DataFrame, logging warnings for any desired
    columns that are not found in the DataFrame.
    This simplifies robust column selection and consistent logging.

    Args:
        df (pd.DataFrame): The input DataFrame.
        desired_columns (List[str]): A list of column names to select.
        logger_instance (Optional[logging.Logger]): An optional logger instance to use for messages.
                                                     If None, the module's default logger is used.

    Returns:
        pd.DataFrame: A new DataFrame containing only the columns that exist in the input
                      DataFrame from the `desired_columns` list.
    """
    current_logger = logger_instance if logger_instance is not None else logger

    existing_columns = [col for col in desired_columns if col in df.columns]
    missing_columns = [col for col in desired_columns if col not in df.columns]

    if missing_columns:
        current_logger.warning(f"The following desired columns were not found in the DataFrame: {missing_columns}. Only existing columns will be selected.")
    current_logger.info(f"Selecting columns: {existing_columns}")

    return df[existing_columns]


def load_dataframe_from_csv(file_path, logger_instance=None, **kwargs):
    # Use the provided logger instance, or create a default one if none is provided
    log = logger_instance if logger_instance else module_logger
    file_path_str = str(file_path)

    log.debug(f"Attempting to load data from CSV: {file_path_str}")

    if not os.path.exists(file_path_str):
        log.error(f"File not found: {file_path_str}")
        return pd.DataFrame() # Return empty DataFrame on error

    # IMPORTANT: Ensure 'logger_instance' is not in kwargs before passing to pd.read_csv.
    kwargs.pop('logger_instance', None) # Defensive pop

    try:
        df = pd.read_csv(file_path_str, **kwargs)
        ## we use **keywordargumments to allow to pass any other arugumment and as many you want 
        log.info(f"Successfully loaded {len(df)} rows from {file_path_str}")
        log.debug(f"CSV data head:\n{df.head()}")
        return df
    except pd.errors.EmptyDataError:
        log.warning(f"CSV file is empty: {file_path_str}. Returning empty DataFrame.")
        return pd.DataFrame()
    except pd.errors.ParserError as e:
        log.error(f"Error parsing CSV file {file_path_str}: {e}")
        log.exception(f"Full traceback for CSV parsing error for {file_path_str}:")
        return pd.DataFrame()
    except Exception as e:
        log.error(f"An unexpected error occurred while reading CSV {file_path_str}: {e}")
        log.exception(f"Full traceback for unexpected CSV read error for {file_path_str}:")
        return pd.DataFrame()


def load_excel_workbook(file_path: Path, logger_instance: logging.Logger = None):
    """
    Loads an Excel workbook and its active worksheet using openpyxl.

    Args:
        file_path (Path): The path to the Excel file.
        logger_instance (logging.Logger, optional): Logger instance for messages.
                                                  If None, uses module_logger.

    Returns:
        Tuple[openpyxl.workbook.workbook.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
            A tuple containing the loaded workbook and its active worksheet.

    Raises:
        FileNotFoundError: If the Excel file does not exist.
        Exception: For any other errors during file loading.
    """
    log = logger_instance if logger_instance is not None else module_logger
    file_path_str = str(file_path)

    log.debug(f"Attempting to load Excel workbook from: {file_path_str}")

    if not Path(file_path_str).exists():
        log.critical(f"Excel file not found: {file_path_str}")
        raise FileNotFoundError(f"Excel file not found: {file_path_str}")

    try:
        wb = load_workbook(file_path_str)
        ws = wb.active
        log.info(f"Successfully loaded Excel workbook '{file_path.name}' and its active worksheet.")
        log.debug(f"Active worksheet name: '{ws.title}'")
        return wb, ws
    except Exception as e:
        log.critical(f"Error loading Excel workbook '{file_path_str}': {e}")
        log.exception(f"Full traceback for Excel workbook loading error for {file_path_str}:")
        raise # Re-raise to be caught by the calling script

def save_dataframe_to_csv(df: pd.DataFrame, file_path: Path, logger_instance: logging.Logger = None, **kwargs: Any) -> None: # <-- CHANGED 'logger' to 'logger_instance'
    """
    Saves a pandas DataFrame to a CSV file, ensuring the output directory exists.

    Parameters:
    - df: The pandas DataFrame to save.
    - file_path: The full path where the CSV file should be saved.
    - logger_instance: Optional logging.Logger instance for logging status/errors. # <-- UPDATED DOCSTRING
    - **kwargs: Additional keyword arguments to pass directly to pandas.DataFrame.to_csv.
                'index=False' is set by default but can be overridden.
    """
    file_path = Path(file_path) # Ensure it's a Path object
    try:
        file_path.parent.mkdir(parents=True, exist_ok=True) # Create parent directories if they don't exist
        
        # Set default for index if not provided, allowing override via kwargs
        if 'index' not in kwargs:
            kwargs['index'] = False
            
        df.to_csv(file_path, **kwargs)
        if logger_instance: # <-- CHANGED 'logger' to 'logger_instance'
            logger_instance.info(f"DataFrame successfully saved to {file_path}") # <-- CHANGED 'logger' to 'logger_instance'
    except Exception as e:
        if logger_instance: # <-- CHANGED 'logger' to 'logger_instance'
            logger_instance.critical(f"Failed to save DataFrame to {file_path}: {e}") # <-- CHANGED 'logger' to 'logger_instance'
        raise


def build_time_series_filepath(
    base_dir: Union[str, Path],
    year: str,
    month: str,
    day: str,
    hour: int,
    filename_prefix: str,
    filename_suffix: str
) -> Path:
    """
    Constructs a file path for time-series data with a YYYY/MM/DD directory structure.
    This simplifies repetitive path construction for dated files.

    Args:
        base_dir (Union[str, Path]): The base directory where the time-series data is stored.
        year (str): The year (e.g., '2023').
        month (str): The month (e.g., '01').
        day (str): The day (e.g., '15').
        hour (int): The hour (0-23).
        filename_prefix (str): Prefix for the filename (e.g., 'Synop_').
        filename_suffix (str): Suffix for the filename (e.g., '.bufr').

    Returns:
        Path: The constructed absolute file path using pathlib.
    """
    base_path = Path(base_dir).resolve()
    # Using f-strings for consistent filename formatting
    file_path = base_path / year / month / day / f"{filename_prefix}{year}{month}{day}{hour:02d}00{filename_suffix}"
    logger.debug(f"Constructed file path: {file_path}")
    return file_path


def save_dataframe_to_excel(
    df: pd.DataFrame,
    file_path: Union[str, Path],
    sheet_name: str = "Sheet1",
    include_header: bool = True,
    include_index: bool = False,
    logger_instance: Optional[logging.Logger] = None,
    **kwargs: Any
) -> None:
    """
    Saves a pandas DataFrame to an Excel file using openpyxl.
    Ensures the parent directory exists.

    Args:
        df (pd.DataFrame): The DataFrame to save.
        file_path (Union[str, Path]): The full path to the output Excel file.
        sheet_name (str): The name of the sheet within the workbook. Defaults to "Sheet1".
        include_header (bool): Whether to include the DataFrame header in the Excel file. Defaults to True.
        include_index (bool): Whether to include the DataFrame index in the Excel file. Defaults to False.
        logger_instance (Optional[logging.Logger]): An optional logger instance to use.
                                                     If None, the module's default logger is used.
        **kwargs: Additional keyword arguments to pass to openpyxl's Workbook.save method.
    """
    current_logger = logger_instance if logger_instance is not None else logger
    ## if there is no logger_instance return use logger module 

    path = Path(file_path)
   ## use Path module to get object directory 

    try:
        # Ensure the parent directory exists
        path.parent.mkdir(parents=True, exist_ok=True)
        ##path.parent figure out the parent folder (the last folder that script is saved )
        ##mkdir (parents= true ) allow the code to creat all the path if any folder does not exist that mean if any parent folder will not exist creat it intel we get the full directory (other wise fi we don't user it mkdir it is allowed to creat one folder at ones)
        ## mkdir(exist_ok=True) the exist_ok = true that mean if the folder is exist don't do any thing and exit mkdir without rise any error

        current_logger.debug(f"Ensured directory exists for Excel file: {path.parent}")

        wb, ws = create_workbook_with_sheet(sheet_name)
        
        for row in dataframe_to_rows(df, index=include_index, header=include_header):
            ws.append(row)
        ## loop through the df dataframe and append (add) that row to worksheet

        wb.save(path)
        ## save the workbook in the path 
        current_logger.info(f"DataFrame successfully saved to Excel file: {path} (Sheet: {sheet_name})")
    except Exception as e:
        current_logger.error(f"Failed to save DataFrame to Excel file '{path}': {e}", exc_info=True)
        raise

# --- Helper function for save_dataframe_to_excel (MUST be defined before save_dataframe_to_excel) ---
def create_workbook_with_sheet(sheet_name: str) -> tuple[Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    """
    Creates a new openpyxl workbook and returns the workbook and the specified sheet.
    If the sheet_name is 'Sheet', it renames the default active sheet.
    Otherwise, it creates a new sheet.
    """
    wb = Workbook()
    ## creat new workbook
    # openpyxl creates a default sheet named 'Sheet'
    
    if sheet_name == "Sheet":
    
        ws = wb.active
        ## activate workbook and it by default creat new worksheet
        ## it does only show first one open workbook
         
        ws.title = sheet_name
        ## set title for worksheet
    else:
        # Remove default 'Sheet' if it's the only one and we're creating a new one
        if len(wb.sheetnames) == 1 and wb.sheetnames[0] == 'Sheet':
            wb.remove(wb.active)
        ws = wb.create_sheet(title=sheet_name)
    return wb, ws


def load_excel_workbook(file_path: Union[str, Path], logger_instance: Optional[logging.Logger] = None) -> Tuple[Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    """
    Loads an Excel workbook and returns the workbook object and its active sheet.
    Includes robust error handling and logging.

    Args:
        file_path (Union[str, Path]): Path to the Excel file.
        logger_instance (Optional[logging.Logger]): Optional logging.Logger instance for messages.

    Returns:
        Tuple[Workbook, openpyxl.worksheet.worksheet.Worksheet]: The loaded Workbook object and its active Worksheet object.

    Raises:
        FileNotFoundError: If the Excel file does not exist.
        Exception: For any other errors during the loading process.
    """
    current_logger = logger_instance if logger_instance is not None else logger
    path = Path(file_path)
    try:
        current_logger.info(f"Attempting to load Excel workbook: {path}")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        current_logger.info(f"Successfully loaded Excel workbook: {path}, active sheet: '{ws.title}'")
        return wb, ws
    except FileNotFoundError:
        current_logger.critical(f"Error: Excel file not found at {path}. Please ensure the file exists.")
        raise
    except Exception as e:
        current_logger.critical(f"Error loading Excel workbook {path}: {e}", exc_info=True)
        raise



def read_bufr_to_dataframe(
    bufr_file_path: Union[str, Path],
    columns: Union[str, List[str]],
    logger_instance: Optional[logging.Logger] = None,
    **pdbufr_kwargs: Any # <--- ADD THIS LINE
) -> Optional[pd.DataFrame]:
    """
    Reads specified columns from a BUFR file into a Pandas DataFrame.

    Args:
        bufr_file_path (Union[str, Path]): Path to the BUFR file.
        columns (Union[str, List[str]]): A single column name or a list of column names
                                         (BUFR descriptor names) to extract.
        logger_instance (Optional[logging.Logger]): An optional logger instance to use for messages.
                                                    If None, the module's default logger is used.
        **pdbufr_kwargs: Additional keyword arguments to pass directly to pdbufr.read_bufr
                         (e.g., 'filters', 'errors', 'encoding'). # <--- ADD THIS DOCSTRING PART
    Returns:
        Optional[pd.DataFrame]: A Pandas DataFrame with the requested columns,
                                or None if the file cannot be processed or no data is found.
    """
    file_path = Path(bufr_file_path).resolve()
    current_logger = logger_instance if logger_instance is not None else logger

    current_logger.info(f"Attempting to read BUFR file: {file_path} for columns: {columns} with kwargs: {pdbufr_kwargs}")
    try:
        df = pdbufr.read_bufr(file_path, columns=columns, **pdbufr_kwargs) # <--- PASS KWARGS HERE
        current_logger.debug(f"Successfully read {len(df)} records from {file_path}.")

        if df.empty:
            current_logger.warning(f"No data records found in {file_path} for columns {columns} with filters {pdbufr_kwargs.get('filters', 'None')}.")
            return None
        
        current_logger.info(f"Successfully loaded BUFR data from {file_path}. DataFrame shape: {df.shape}")
        return df

    except Exception as e:
        current_logger.error(f"An error occurred while reading BUFR file {file_path}: {e}", exc_info=True)
        return None
# ... (rest of your pytaps/data_utils.py content) ...